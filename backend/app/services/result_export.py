"""Phase 34 — query-result serialisers for the CSV / Excel / JSON
download endpoints.

The endpoint shape is intentionally generic: each format takes the
same ``(columns, rows)`` tuple as input and returns ``bytes`` plus a
MIME type. The chat API caches result rows on ``query_history``
(see migration 0022) so the endpoint never re-runs the user's query.

Excel uses ``openpyxl`` which is already in the project's deps
(pulled in by the Phase 14 doc harvester for XLSX text extraction).
CSV uses stdlib ``csv``. JSON uses stdlib ``json``.

Cell coercion: pgvector, datetime, Decimal, UUID, bytes — none of
these survive a raw ``json.dumps`` or a naive XLSX cell write. We
normalise to:
  * datetime → ISO 8601 string
  * Decimal  → float (lossy but matches what the UI already shows)
  * UUID     → string
  * bytes    → ``"<binary, N bytes>"`` placeholder
  * other    → ``repr(x)`` as the safety net
"""
from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any
from uuid import UUID


def _coerce_cell(v: Any) -> Any:
    """Normalise one cell value for CSV / XLSX / JSON output. Returns
    a JSON-safe primitive."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    if isinstance(v, Decimal):
        # JSON has no Decimal — lossy but matches UI rendering.
        return float(v)
    if isinstance(v, UUID):
        return str(v)
    if isinstance(v, (bytes, bytearray, memoryview)):
        return f"<binary, {len(v)} bytes>"
    if isinstance(v, (list, tuple)):
        return [_coerce_cell(x) for x in v]
    if isinstance(v, dict):
        return {str(k): _coerce_cell(val) for k, val in v.items()}
    # Last resort. repr keeps a stable, debuggable representation.
    return repr(v)


def _coerce_rows(rows: list[list[Any]]) -> list[list[Any]]:
    return [[_coerce_cell(c) for c in row] for row in rows]


# ── CSV ──────────────────────────────────────────────────────────


def to_csv(columns: list[str], rows: list[list[Any]]) -> bytes:
    """RFC-4180-flavoured CSV with UTF-8 BOM so Excel on Windows
    autodetects the encoding instead of mangling Cyrillic / Uzbek
    Latin diacritics. ``\r\n`` line terminator for the same reason.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow(columns)
    for row in _coerce_rows(rows):
        writer.writerow(row)
    body = buf.getvalue().encode("utf-8")
    # UTF-8 BOM for Excel autodetect.
    return b"\xef\xbb\xbf" + body


# ── JSON ─────────────────────────────────────────────────────────


def to_json(columns: list[str], rows: list[list[Any]]) -> bytes:
    """Row-oriented JSON — one object per row. Preferred shape for
    programmatic consumers (Python pandas, Node fetch, jq pipelines).
    Pretty-printed with indent=2 for grep-ability."""
    coerced = _coerce_rows(rows)
    payload = [dict(zip(columns, row)) for row in coerced]
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


# ── Excel ────────────────────────────────────────────────────────


def to_xlsx(columns: list[str], rows: list[list[Any]]) -> bytes:
    """Single-sheet XLSX. Bold header row, frozen first row, column
    widths auto-sized to the longest visible string (capped at 60
    chars so a giant cell doesn't push the sheet off-screen).

    openpyxl-only — no pandas dependency. We write to an in-memory
    BytesIO and return raw bytes; the endpoint hands these to
    StreamingResponse.
    """
    # Local import keeps openpyxl off the import path for installs
    # that don't enable the harvester or the export feature.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F2937", end_color="1F2937", fill_type="solid"
    )
    for ci, name in enumerate(columns, start=1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")

    # Rows
    coerced = _coerce_rows(rows)
    for ri, row in enumerate(coerced, start=2):
        for ci, val in enumerate(row, start=1):
            # XLSX cells can hold strings, numbers, bools, datetimes
            # natively. Lists / dicts must become JSON strings —
            # otherwise openpyxl raises TypeError.
            if isinstance(val, (list, dict)):
                ws.cell(row=ri, column=ci, value=json.dumps(
                    val, ensure_ascii=False
                ))
            else:
                ws.cell(row=ri, column=ci, value=val)

    # Column widths — measure header + first 200 rows. Cheap and
    # bounded; full-scan on a 10k-row sheet would be wasteful.
    widths: dict[int, int] = {}
    for ci, name in enumerate(columns, start=1):
        widths[ci] = max(widths.get(ci, 0), len(str(name)))
    sample = coerced[:200]
    for row in sample:
        for ci, val in enumerate(row, start=1):
            w = len(str(val)) if val is not None else 0
            if w > widths.get(ci, 0):
                widths[ci] = w
    for ci, w in widths.items():
        # +2 for padding; clamp to 60 so a giant URL doesn't break
        # the layout.
        ws.column_dimensions[get_column_letter(ci)].width = min(
            max(w + 2, 8), 60
        )

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Capture / cap helpers ────────────────────────────────────────


def captured_rows_for_export(
    result_set: Any,
    *,
    max_rows: int,
    max_bytes: int,
) -> tuple[list[str] | None, list[list[Any]] | None]:
    """Decide whether a ResultSet should be persisted for export.

    Drops the cache (returns ``(None, None)``) when:
      * the result has more rows than ``max_rows``
      * the JSON-serialised payload would exceed ``max_bytes``
      * the result is missing entirely

    The chat audit row (``query_history``) still records the dialect /
    SQL / row_count even when we drop the rows — only the export
    endpoint is degraded.
    """
    if result_set is None:
        return None, None
    columns = list(getattr(result_set, "columns", []) or [])
    rows = list(getattr(result_set, "rows", []) or [])
    if not columns:
        return None, None
    if len(rows) > max_rows:
        return None, None
    # Cheap byte budget: serialise the coerced rows + columns and
    # compare lengths. JSONB on Postgres has effectively no limit but
    # we want predictable storage cost.
    try:
        coerced = _coerce_rows(rows)
        payload = json.dumps(
            {"columns": columns, "rows": coerced},
            ensure_ascii=False,
        )
    except (TypeError, ValueError):
        # Some opaque cell type that doesn't survive coercion — skip
        # the cache rather than corrupt the audit row.
        return None, None
    if len(payload.encode("utf-8")) > max_bytes:
        return None, None
    return columns, coerced


__all__ = [
    "to_csv",
    "to_json",
    "to_xlsx",
    "captured_rows_for_export",
]
