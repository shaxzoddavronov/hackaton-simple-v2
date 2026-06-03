"""Phase 34 — query-result export serialisers."""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from decimal import Decimal
from types import SimpleNamespace
from uuid import UUID

import pytest

from app.services.result_export import (
    _coerce_cell,
    captured_rows_for_export,
    to_csv,
    to_json,
    to_xlsx,
)


# ── cell coercion ────────────────────────────────────────────────


def test_coerce_passes_through_primitives() -> None:
    assert _coerce_cell(None) is None
    assert _coerce_cell(42) == 42
    assert _coerce_cell(3.14) == 3.14
    assert _coerce_cell(True) is True
    assert _coerce_cell("hello") == "hello"


def test_coerce_datetime_to_iso() -> None:
    dt = datetime(2026, 6, 3, 12, 0, 0, tzinfo=timezone.utc)
    assert _coerce_cell(dt) == "2026-06-03T12:00:00+00:00"


def test_coerce_decimal_to_float() -> None:
    assert _coerce_cell(Decimal("3.14")) == pytest.approx(3.14)


def test_coerce_uuid_to_string() -> None:
    uid = UUID("00000000-0000-0000-0000-000000000001")
    assert _coerce_cell(uid) == str(uid)


def test_coerce_bytes_to_placeholder() -> None:
    out = _coerce_cell(b"abcdef")
    assert "binary" in out
    assert "6 bytes" in out


def test_coerce_recurses_into_lists_and_dicts() -> None:
    assert _coerce_cell([1, 2, Decimal("3.5")]) == [1, 2, 3.5]
    assert _coerce_cell({"a": Decimal("1")}) == {"a": 1.0}


# ── CSV ──────────────────────────────────────────────────────────


def test_to_csv_has_bom_and_crlf() -> None:
    body = to_csv(["a", "b"], [[1, 2]])
    # UTF-8 BOM for Excel auto-detect
    assert body.startswith(b"\xef\xbb\xbf")
    text = body[3:].decode("utf-8")
    assert "\r\n" in text


def test_to_csv_quotes_commas_and_newlines() -> None:
    body = to_csv(["x"], [["a, b"], ["line1\nline2"]])
    text = body[3:].decode("utf-8")
    reader = csv.reader(io.StringIO(text))
    assert next(reader) == ["x"]
    assert next(reader) == ["a, b"]
    assert next(reader) == ["line1\nline2"]


def test_to_csv_handles_multilingual_strings() -> None:
    body = to_csv(
        ["uz", "ru", "en"],
        [["Salom", "Привет", "Hello"]],
    )
    text = body[3:].decode("utf-8")
    assert "Salom" in text
    assert "Привет" in text


def test_to_csv_coerces_datetime() -> None:
    body = to_csv(
        ["t"],
        [[datetime(2026, 6, 3, tzinfo=timezone.utc)]],
    )
    text = body[3:].decode("utf-8")
    assert "2026-06-03" in text


def test_to_csv_empty_rows_still_emits_header() -> None:
    body = to_csv(["a", "b"], [])
    text = body[3:].decode("utf-8")
    assert text.startswith("a,b")


# ── JSON ─────────────────────────────────────────────────────────


def test_to_json_emits_row_objects() -> None:
    body = to_json(["id", "name"], [[1, "alice"], [2, "bob"]])
    parsed = json.loads(body.decode("utf-8"))
    assert parsed == [
        {"id": 1, "name": "alice"},
        {"id": 2, "name": "bob"},
    ]


def test_to_json_coerces_decimal_and_datetime() -> None:
    body = to_json(
        ["amount", "when"],
        [[Decimal("99.99"), datetime(2026, 6, 3, tzinfo=timezone.utc)]],
    )
    parsed = json.loads(body.decode("utf-8"))
    assert parsed[0]["amount"] == 99.99
    assert parsed[0]["when"].startswith("2026-06-03")


def test_to_json_preserves_unicode() -> None:
    body = to_json(["name"], [["Тошкент"]])
    text = body.decode("utf-8")
    # ensure_ascii=False keeps the Cyrillic intact
    assert "Тошкент" in text


# ── XLSX ─────────────────────────────────────────────────────────


def test_to_xlsx_is_valid_workbook() -> None:
    body = to_xlsx(
        ["id", "name"],
        [[1, "a"], [2, "b"]],
    )
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    assert ws["A1"].value == "id"
    assert ws["B1"].value == "name"
    assert ws["A2"].value == 1
    assert ws["B2"].value == "a"
    assert ws["B3"].value == "b"


def test_to_xlsx_freezes_header_row() -> None:
    body = to_xlsx(["a"], [[1]])
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(body))
    assert wb.active.freeze_panes == "A2"


def test_to_xlsx_styles_header() -> None:
    body = to_xlsx(["x"], [])
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(body))
    header_cell = wb.active["A1"]
    assert header_cell.font.bold is True


def test_to_xlsx_handles_nested_cells_as_json() -> None:
    body = to_xlsx(
        ["tags"],
        [[["red", "blue"]], [{"k": "v"}]],
    )
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    # Nested types serialised to JSON strings before write.
    a2 = ws["A2"].value
    a3 = ws["A3"].value
    assert isinstance(a2, str)
    assert "red" in a2
    assert isinstance(a3, str)
    assert '"k": "v"' in a3


def test_to_xlsx_coerces_datetime_to_native() -> None:
    body = to_xlsx(
        ["t"],
        [[datetime(2026, 6, 3, tzinfo=timezone.utc)]],
    )
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(body))
    val = wb.active["A2"].value
    # _coerce_cell turned datetime into ISO string, so XLSX stores a string.
    assert isinstance(val, str)
    assert val.startswith("2026-06-03")


# ── captured_rows_for_export ─────────────────────────────────────


def _rs(columns: list[str], rows: list[list]) -> SimpleNamespace:
    return SimpleNamespace(columns=columns, rows=rows)


def test_capture_returns_columns_and_rows_when_in_budget() -> None:
    cols, rows = captured_rows_for_export(
        _rs(["a", "b"], [[1, 2], [3, 4]]),
        max_rows=10,
        max_bytes=1024,
    )
    assert cols == ["a", "b"]
    assert rows == [[1, 2], [3, 4]]


def test_capture_drops_when_row_cap_exceeded() -> None:
    cols, rows = captured_rows_for_export(
        _rs(["a"], [[i] for i in range(100)]),
        max_rows=10,
        max_bytes=1024 * 1024,
    )
    assert cols is None and rows is None


def test_capture_drops_when_byte_cap_exceeded() -> None:
    long = "x" * 5000
    cols, rows = captured_rows_for_export(
        _rs(["s"], [[long] for _ in range(50)]),
        max_rows=10_000,
        max_bytes=1024,  # 1 KiB — way under
    )
    assert cols is None and rows is None


def test_capture_handles_none_result() -> None:
    cols, rows = captured_rows_for_export(
        None, max_rows=10, max_bytes=1024
    )
    assert cols is None and rows is None


def test_capture_handles_empty_columns() -> None:
    cols, rows = captured_rows_for_export(
        _rs([], []), max_rows=10, max_bytes=1024
    )
    assert cols is None and rows is None


def test_capture_coerces_datetime_inside_payload() -> None:
    dt = datetime(2026, 6, 3, tzinfo=timezone.utc)
    cols, rows = captured_rows_for_export(
        _rs(["t"], [[dt]]), max_rows=10, max_bytes=1024
    )
    assert cols == ["t"]
    # Coerced to ISO string, not raw datetime.
    assert rows == [["2026-06-03T00:00:00+00:00"]]
